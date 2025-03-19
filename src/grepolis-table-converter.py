import pandas as pd
import os
import sys
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

COLOR_MAPPING = {
    "FFFF00": "player",
    "92D050": "town",
    "00B0F0": "ally"
}


def find_table_bounds(df):
    non_empty_rows = df.dropna(how='all').index
    non_empty_cols = df.dropna(axis=1, how='all').columns

    if non_empty_rows.empty or non_empty_cols.empty:
        return None, None, None, None

    return non_empty_rows[0] + 1, non_empty_rows[-1] + 1, non_empty_cols[0] + 1, non_empty_cols[-1] + 1


def excel_to_bbcode(file_path, start_row, end_row, start_col, end_col, enable_headline):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        bbcode = "[table]\n"

        # Process headline row if enabled.
        if enable_headline:
            header_cells = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=start_row, column=col)
                text = str(cell.value) if cell.value is not None else ""
                # Check the cell's fill color.
                fill_color = cell.fill.start_color.rgb
                if fill_color and len(fill_color) == 8:
                    # Remove the alpha channel if present.
                    fill_color = fill_color[2:]
                if fill_color in COLOR_MAPPING:
                    tag = COLOR_MAPPING[fill_color]
                    text = f"[{tag}]{text}[/{tag}]"
                header_cells.append(text)
            bbcode += "[**]" + "[||]".join(header_cells) + "[/**]\n"
            start_row += 1  # Skip the headline row for the rest of the table.

        # Process the remaining rows.
        for row in range(start_row, end_row + 1):
            row_cells = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                text = str(cell.value) if cell.value is not None else ""
                fill_color = cell.fill.start_color.rgb
                if fill_color and len(fill_color) == 8:
                    fill_color = fill_color[2:]
                if fill_color in COLOR_MAPPING:
                    tag = COLOR_MAPPING[fill_color]
                    text = f"[{tag}]{text}[/{tag}]"
                row_cells.append(text)
            bbcode += "[*]" + "[|]".join(row_cells) + "[/*]\n"

        bbcode += "[/table]"
        return bbcode
    except Exception as e:
        return f"Error processing file: {e}"


def open_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        file_path_var.set(file_path)
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        start, end, left, right = find_table_bounds(df)
        if start:
            table_size_label.config(text=f"Detected Table Size: Rows {start}-{end}, Columns {left}-{right}")
            original_start_row.set(start)
            original_end_row.set(end)
            original_start_col.set(left)
            original_end_col.set(right)
            reset_ranges()


def reset_ranges():
    start_row_entry.delete(0, tk.END)
    start_row_entry.insert(0, original_start_row.get())
    end_row_entry.delete(0, tk.END)
    end_row_entry.insert(0, original_end_row.get())
    start_col_entry.delete(0, tk.END)
    start_col_entry.insert(0, original_start_col.get())
    end_col_entry.delete(0, tk.END)
    end_col_entry.insert(0, original_end_col.get())


def convert_file():
    file_path = file_path_var.get()
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file first.")
        return

    bbcode_output = excel_to_bbcode(
        file_path,
        int(start_row_entry.get()),
        int(end_row_entry.get()),
        int(start_col_entry.get()),
        int(end_col_entry.get()),
        headline_var.get()
    )
    text_output.delete("1.0", tk.END)
    text_output.insert(tk.END, bbcode_output)

    if save_txt_var.get():
        output_file = os.path.splitext(file_path)[0] + ".txt"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(bbcode_output)
        messagebox.showinfo("Success", f"BBCode saved to {output_file}")


def copy_to_clipboard():
    root.clipboard_clear()
    root.clipboard_append(text_output.get("1.0", tk.END).strip())
    root.update()
    messagebox.showinfo("Success", "BBCode copied to clipboard!")


def create_gui():
    global file_path_var, start_row_entry, end_row_entry, start_col_entry, end_col_entry, text_output, headline_var, save_txt_var, table_size_label
    global original_start_row, original_end_row, original_start_col, original_end_col, root

    root = tk.Tk()
    # Set custom icon if available.
    icon_path = os.path.join("resources", "grepolis_excel_icon_v2.ico")
    if os.path.exists(icon_path):
        try:
            root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Unable to set icon: {e}")
    root.title("Excel to Grepolis BB Code Converter")
    frame = tk.Frame(root, padx=10, pady=10)
    frame.pack()

    file_frame = tk.Frame(frame)
    file_frame.pack(fill="x")
    file_path_var = tk.StringVar(value="No file selected")

    tk.Button(file_frame, text="Select Excel File", command=open_file).pack(side=tk.LEFT, padx=(0, 10))
    tk.Label(file_frame, textvariable=file_path_var, width=50, anchor="w").pack(side=tk.LEFT)

    table_size_label = tk.Label(frame, text="Detected Table Size: ")
    table_size_label.pack()

    options_frame = tk.LabelFrame(frame, text="Options", padx=10, pady=10)
    options_frame.pack(fill="x")

    original_start_row = tk.IntVar()
    original_end_row = tk.IntVar()
    original_start_col = tk.IntVar()
    original_end_col = tk.IntVar()

    tk.Label(options_frame, text="Start Row:").grid(row=0, column=0)
    start_row_entry = tk.Entry(options_frame, width=5)
    start_row_entry.grid(row=0, column=1)

    tk.Label(options_frame, text="End Row:").grid(row=0, column=2)
    end_row_entry = tk.Entry(options_frame, width=5)
    end_row_entry.grid(row=0, column=3)

    tk.Label(options_frame, text="Start Column:").grid(row=1, column=0)
    start_col_entry = tk.Entry(options_frame, width=5)
    start_col_entry.grid(row=1, column=1)

    tk.Label(options_frame, text="End Column:").grid(row=1, column=2)
    end_col_entry = tk.Entry(options_frame, width=5)
    end_col_entry.grid(row=1, column=3)

    tk.Button(options_frame, text="Reset", command=reset_ranges).grid(row=0, column=4, rowspan=2)

    headline_var = tk.BooleanVar(value=True)
    tk.Checkbutton(options_frame, text="Enable Headline", variable=headline_var).grid(row=2, column=0, columnspan=2)

    btn_frame = tk.Frame(frame)
    btn_frame.pack(fill="x")
    save_txt_var = tk.BooleanVar(value=True)
    tk.Button(btn_frame, text="Convert to BBCode", command=convert_file).pack(side=tk.LEFT, padx=5, pady=5)
    tk.Checkbutton(btn_frame, text="Save as .txt", variable=save_txt_var).pack(side=tk.LEFT)

    tk.Button(btn_frame, text="Copy to Clipboard", command=copy_to_clipboard).pack(side=tk.RIGHT, padx=5, pady=5)

    text_output = scrolledtext.ScrolledText(frame, height=15, width=80)
    text_output.pack()

    root.mainloop()


if __name__ == "__main__":
    create_gui()
